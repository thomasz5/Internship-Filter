#!/usr/bin/env python3
"""
Excel Integration for UW Internship Finder
Automatically updates live Excel spreadsheet with new opportunities
"""

import sqlite3
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import os
import logging
from config import Config
from aws_monitor import EC2Monitor

class ExcelIntegration:
    def __init__(self):
        self.config = Config()
        self.excel_file = "UW_Internship_Tracker.xlsx"
        self.logger = logging.getLogger(__name__)
        self.ec2_monitor = EC2Monitor()
        
    def create_or_update_excel(self):
        """Create or update the Excel workbook with latest data"""
        try:
            # Get data from database
            internships_df, alumni_df, opportunities_df = self._get_data_from_db()
            
            # Create or load workbook
            if os.path.exists(self.excel_file):
                wb = load_workbook(self.excel_file)
            else:
                wb = Workbook()
                # Remove default sheet
                if 'Sheet' in wb.sheetnames:
                    wb.remove(wb['Sheet'])
            
            # Create/update worksheets
            self._create_opportunities_sheet(wb, opportunities_df)
            self._create_internships_sheet(wb, internships_df)
            self._create_alumni_sheet(wb, alumni_df)
            self._create_aws_status_sheet(wb)
            self._create_summary_sheet(wb, internships_df, alumni_df)
            
            # Save workbook
            wb.save(self.excel_file)
            
            print(f"Excel file updated: {self.excel_file}")
            print(f"   • {len(internships_df)} internship opportunities")
            print(f"   • {len(alumni_df)} UW alumni profiles")
            print(f"   • {len(opportunities_df)} combined opportunities")
            
            # Try to open Excel if on macOS
            self._open_excel_if_possible()
            
            return True
            
        except Exception as e:
            self.logger.error(f"Error updating Excel file: {e}")
            return False
    
    def _get_data_from_db(self):
        """Retrieve data from SQLite database"""
        conn = sqlite3.connect(self.config.DATABASE_PATH)
        
        # Get internships
        internships_df = pd.read_sql_query('''
            SELECT 
                company,
                role,
                location,
                application_link,
                source_repo,
                discovered_date,
                CASE 
                    WHEN discovered_date > datetime('now', '-1 day') THEN 'NEW!'
                    WHEN discovered_date > datetime('now', '-7 days') THEN 'This Week'
                    ELSE 'Older'
                END as freshness
            FROM internships
            ORDER BY discovered_date DESC
        ''', conn)
        
        # Get alumni
        alumni_df = pd.read_sql_query('''
            SELECT 
                name,
                title,
                company,
                linkedin_url,
                discovered_date,
                CASE 
                    WHEN discovered_date > datetime('now', '-1 day') THEN 'NEW!'
                    WHEN discovered_date > datetime('now', '-7 days') THEN 'This Week'
                    ELSE 'Older'
                END as freshness
            FROM profiles
            ORDER BY discovered_date DESC
        ''', conn)
        
        # Get combined opportunities (internships + matching alumni)
        opportunities_df = pd.read_sql_query('''
            SELECT 
                i.company,
                i.role as "Internship Role",
                i.location as "Internship Location",
                i.application_link as "Application Link",
                i.discovered_date as "Internship Found",
                p.name as "UW Alumni Name",
                p.title as "Alumni Title",
                p.linkedin_url as "Alumni LinkedIn",
                p.discovered_date as "Alumni Found",
                CASE 
                    WHEN i.discovered_date > datetime('now', '-1 day') OR p.discovered_date > datetime('now', '-1 day') THEN 'NEW!'
                    WHEN i.discovered_date > datetime('now', '-7 days') OR p.discovered_date > datetime('now', '-7 days') THEN 'This Week'
                    ELSE 'Older'
                END as "Status"
            FROM internships i
            LEFT JOIN profiles p ON i.company = p.company
            ORDER BY i.discovered_date DESC, p.discovered_date DESC
        ''', conn)
        
        conn.close()
        return internships_df, alumni_df, opportunities_df
    
    def _create_opportunities_sheet(self, wb, df):
        """Create the main opportunities overview sheet"""
        sheet_name = "🎯 Opportunities"
        
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        
        ws = wb.create_sheet(sheet_name, 0)  # Make it first sheet
        
        if df.empty:
            ws['A1'] = "No opportunities found yet. The system will update this automatically!"
            ws['A1'].font = Font(size=14, bold=True)
            return
        
        # Add headers
        headers = ["Company", "Internship Role", "Location", "Application Link", 
                  "UW Alumni Name", "Alumni Title", "Alumni LinkedIn", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add data
        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            ws.cell(row=row_idx, column=1, value=row.get('company', ''))
            ws.cell(row=row_idx, column=2, value=row.get('Internship Role', ''))
            ws.cell(row=row_idx, column=3, value=row.get('Internship Location', ''))
            
            # Make application link clickable
            app_link = row.get('Application Link', '')
            if app_link and app_link.startswith('http'):
                ws.cell(row=row_idx, column=4).hyperlink = app_link
                ws.cell(row=row_idx, column=4).value = "Apply Here"
                ws.cell(row=row_idx, column=4).font = Font(color="0000FF", underline="single")
            else:
                ws.cell(row=row_idx, column=4, value=app_link)
            
            ws.cell(row=row_idx, column=5, value=row.get('UW Alumni Name', ''))
            ws.cell(row=row_idx, column=6, value=row.get('Alumni Title', ''))
            
            # Make LinkedIn profile clickable
            linkedin_url = row.get('Alumni LinkedIn', '')
            if linkedin_url and linkedin_url.startswith('http'):
                ws.cell(row=row_idx, column=7).hyperlink = linkedin_url
                ws.cell(row=row_idx, column=7).value = "LinkedIn Profile"
                ws.cell(row=row_idx, column=7).font = Font(color="0000FF", underline="single")
            else:
                ws.cell(row=row_idx, column=7, value=linkedin_url)
            
            # Status with color coding
            status = row.get('Status', '')
            status_cell = ws.cell(row=row_idx, column=8, value=status)
            if status == 'NEW!':
                status_cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                status_cell.font = Font(bold=True)
            elif status == 'This Week':
                status_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add auto-filter
        ws.auto_filter.ref = f"A1:H{len(df) + 1}"
    
    def _create_internships_sheet(self, wb, df):
        """Create internships-only sheet"""
        sheet_name = "Internships"
        
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        
        ws = wb.create_sheet(sheet_name)
        
        if df.empty:
            ws['A1'] = "No internships found yet."
            return
        
        # Convert dataframe to rows and add to worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Format header row
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add auto-filter
        max_col_letter = ws.cell(row=1, column=ws.max_column).column_letter
        ws.auto_filter.ref = f"A1:{max_col_letter}{ws.max_row}"
    
    def _create_alumni_sheet(self, wb, df):
        """Create UW alumni sheet"""
        sheet_name = "UW Alumni"
        
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        
        ws = wb.create_sheet(sheet_name)
        
        if df.empty:
            ws['A1'] = "No UW alumni found yet."
            return
        
        # Convert dataframe to rows and add to worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Format header row
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4B0082", end_color="4B0082", fill_type="solid")  # Purple for UW
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add auto-filter
        max_col_letter = ws.cell(row=1, column=ws.max_column).column_letter
        ws.auto_filter.ref = f"A1:{max_col_letter}{ws.max_row}"
    
    def _create_aws_status_sheet(self, wb):
        """Create AWS EC2 status monitoring sheet"""
        sheet_name = "☁️ AWS Status"
        
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        
        ws = wb.create_sheet(sheet_name)
        
        # Title
        ws['A1'] = "AWS EC2 Instance Status"
        ws['A1'].font = Font(size=16, bold=True, color="FF6600")
        ws.merge_cells('A1:G1')
        
        # Get EC2 status
        try:
            print("   🔍 Checking AWS EC2 status...")
            ec2_status = self.ec2_monitor.check_all_regions()
            
            if 'error' in ec2_status:
                # Show error information
                ws['A3'] = "❌ Error connecting to AWS"
                ws['A3'].font = Font(bold=True, color="FF0000")
                ws['A4'] = ec2_status['error']
                ws['A6'] = "To fix this issue:"
                ws['A6'].font = Font(bold=True)
                
                instructions = [
                    "1. Install AWS CLI: pip install awscli",
                    "2. Configure credentials: aws configure",
                    "3. Or set environment variables:",
                    "   AWS_ACCESS_KEY_ID=your-key-id",
                    "   AWS_SECRET_ACCESS_KEY=your-secret-key",
                    "4. Alternatively, run manually:",
                    "   python3 src/aws_monitor.py"
                ]
                
                for i, instruction in enumerate(instructions, 7):
                    ws[f'A{i}'] = instruction
                    if instruction.startswith('   '):
                        ws[f'A{i}'].font = Font(italic=True, color="666666")
                
                return
            
            # Summary information
            ws['A3'] = f"Last Checked: {ec2_status['timestamp']}"
            ws['A3'].font = Font(italic=True)
            
            ws['A5'] = "Summary:"
            ws['A5'].font = Font(bold=True)
            
            ws['A6'] = f"Total Regions Checked: {ec2_status['total_regions_checked']}"
            ws['A7'] = f"Regions with Instances: {ec2_status['regions_with_instances']}"
            ws['A8'] = f"Total Instances: {ec2_status['total_instances']}"
            ws['A9'] = f"Running: {ec2_status['total_running']}"
            ws['A9'].fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            ws['A10'] = f"Stopped: {ec2_status['total_stopped']}"
            if ec2_status['total_stopped'] > 0:
                ws['A10'].fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            
            # Instance details header
            current_row = 12
            ws[f'A{current_row}'] = "Instance Details:"
            ws[f'A{current_row}'].font = Font(bold=True)
            current_row += 1
            
            # Headers for instance table
            headers = ["Region", "Name", "Instance ID", "Type", "State", "Public IP", "Launch Time"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            current_row += 1
            
            # Add instance data
            for region_data in ec2_status.get('regions', []):
                region_name = region_data['region']
                for instance in region_data['instances']:
                    ws.cell(row=current_row, column=1, value=region_name)
                    ws.cell(row=current_row, column=2, value=instance['name'])
                    ws.cell(row=current_row, column=3, value=instance['instance_id'])
                    ws.cell(row=current_row, column=4, value=instance['instance_type'])
                    
                    # State with color coding
                    state_cell = ws.cell(row=current_row, column=5, value=instance['state'])
                    if instance['state'] == 'running':
                        state_cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                        state_cell.font = Font(bold=True)
                    elif instance['state'] == 'stopped':
                        state_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                    elif instance['state'] in ['pending', 'stopping', 'starting']:
                        state_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    
                    ws.cell(row=current_row, column=6, value=instance['public_ip'])
                    ws.cell(row=current_row, column=7, value=instance['launch_time'])
                    
                    current_row += 1
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add auto-filter to instance table
            if current_row > 14:  # If we have instance data
                ws.auto_filter.ref = f"A13:G{current_row-1}"
            
            # Cost estimation section
            if ec2_status['total_instances'] > 0:
                cost_row = current_row + 2
                ws[f'A{cost_row}'] = "💰 Cost Estimates (30 days):"
                ws[f'A{cost_row}'].font = Font(bold=True)
                
                try:
                    cost_data = self.ec2_monitor.get_instance_costs()
                    if 'error' not in cost_data:
                        cost_row += 1
                        ws[f'A{cost_row}'] = f"Total Estimated: ${cost_data['total_estimated_cost']}"
                        ws[f'A{cost_row}'].font = Font(bold=True, color="FF6600")
                        
                        cost_row += 1
                        ws[f'A{cost_row}'] = "(Estimates based on standard pricing - check AWS billing for actual costs)"
                        ws[f'A{cost_row}'].font = Font(italic=True, size=10)
                except:
                    cost_row += 1
                    ws[f'A{cost_row}'] = "Cost estimation unavailable"
            
        except Exception as e:
            self.logger.error(f"Error getting AWS status: {e}")
            ws['A3'] = f"❌ Error retrieving AWS status: {str(e)}"
            ws['A3'].font = Font(color="FF0000")
            ws['A5'] = "Try running: python3 src/aws_monitor.py"
            ws['A5'].font = Font(italic=True)
    
    def _create_summary_sheet(self, wb, internships_df, alumni_df):
        """Create summary dashboard sheet"""
        sheet_name = "Dashboard"
        
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        
        ws = wb.create_sheet(sheet_name)
        
        # Title
        ws['A1'] = "UW Internship Finder - Dashboard"
        ws['A1'].font = Font(size=20, bold=True, color="2F75B5")
        ws.merge_cells('A1:D1')
        
        # Last updated
        ws['A3'] = f"Last Updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A3'].font = Font(italic=True)
        
        # Statistics
        ws['A5'] = "Summary Statistics"
        ws['A5'].font = Font(size=14, bold=True)
        
        ws['A7'] = "Total Internships Found:"
        ws['B7'] = len(internships_df)
        ws['B7'].font = Font(bold=True, size=12)
        
        ws['A8'] = "Total UW Alumni Found:"
        ws['B8'] = len(alumni_df)
        ws['B8'].font = Font(bold=True, size=12)
        
        # Recent activity
        if not internships_df.empty:
            new_internships = len(internships_df[internships_df['freshness'] == 'NEW!'])
            week_internships = len(internships_df[internships_df['freshness'] == 'This Week'])
            
            ws['A10'] = "New Today:"
            ws['B10'] = f"{new_internships} internships"
            if new_internships > 0:
                ws['B10'].fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            
            ws['A11'] = "This Week:"
            ws['B11'] = f"{week_internships} internships"
            if week_internships > 0:
                ws['B11'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Instructions
        ws['A14'] = "🎯 How to Use This Tracker"
        ws['A14'].font = Font(size=14, bold=True)
        
        instructions = [
            "1. Check the 'Opportunities' tab for the best overview",
            "2. Green = NEW opportunities found today", 
            "3. Yellow = Found this week",
            "4. Click blue links to apply or view LinkedIn profiles",
            "5. Use filters to search by company, location, etc.",
            "6. This file updates automatically every 12 hours"
        ]
        
        for i, instruction in enumerate(instructions, 16):
            ws[f'A{i}'] = instruction
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 25
    
    def _open_excel_if_possible(self):
        """Try to open Excel file automatically (macOS)"""
        try:
            import subprocess
            # Try to open with Excel specifically
            subprocess.run(['open', '-a', 'Microsoft Excel', self.excel_file], 
                         check=False, capture_output=True)
        except:
            try:
                # Fallback to default application
                import subprocess
                subprocess.run(['open', self.excel_file], 
                             check=False, capture_output=True)
            except:
                # Silent fail - user can open manually
                pass
    
    def add_notification_to_excel(self, message: str):
        """Add a notification/alert to the Excel file"""
        try:
            if not os.path.exists(self.excel_file):
                return
            
            wb = load_workbook(self.excel_file)
            
            # Add to dashboard
            if "Dashboard" in wb.sheetnames:
                ws = wb["Dashboard"]
                
                # Find next available row for notifications
                notification_row = 25
                while ws[f'A{notification_row}'].value:
                    notification_row += 1
                
                if notification_row == 25:
                    ws['A23'] = "🔔 Recent Alerts"
                    ws['A23'].font = Font(size=14, bold=True)
                
                ws[f'A{notification_row}'] = f"{datetime.now().strftime('%H:%M')} - {message}"
                ws[f'A{notification_row}'].fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            
            wb.save(self.excel_file)
            
        except Exception as e:
            self.logger.debug(f"Could not add notification to Excel: {e}")

if __name__ == "__main__":
    excel = ExcelIntegration()
    excel.create_or_update_excel() 